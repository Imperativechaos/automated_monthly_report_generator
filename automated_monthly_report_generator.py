import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime

def generate_monthly_report(file_paths, output_file="monthly_report.xlsx"):
    """
    Merges multiple Excel/CSV files, calculates summary stats,
    and creates a formatted summary report.
    """
    try:
        combined_df = pd.DataFrame()
        
        for file in file_paths:
            if file.endswith('.csv'):
                df = pd.read_csv(file)
            else:
                df = pd.read_excel(file)
            combined_df = pd.concat([combined_df, df], ignore_index=True)
        
        # Example summary: group by category and sum sales
        if 'Category' in combined_df.columns and 'Sales' in combined_df.columns:
            summary = combined_df.groupby('Category')['Sales'].agg(['sum', 'mean', 'count']).reset_index()
            summary.columns = ['Category', 'Total Sales', 'Average Sales', 'Transaction Count']
        else:
            summary = combined_df.describe()  # Fallback generic stats
        
        # Create new workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Summary Report"
        
        # Add title
        ws['A1'] = f"Monthly Report - {datetime.now().strftime('%B %Y')}"
        ws['A1'].font = Font(bold=True, size=14)
        
        # Write summary DataFrame
        for r_idx, row in enumerate(dataframe_to_rows(summary, index=False, header=True), 3):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
        
        wb.save(output_file)
        print(f"Report generated: {output_file}")
        return output_file
    
    except Exception as e:
        print(f"Report generation failed: {e}")
        return None

# Example
# files = ["sales_jan.xlsx", "sales_feb.csv"]
# generate_monthly_report(files)
